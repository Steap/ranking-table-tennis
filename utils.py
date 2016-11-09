import csv
import models
import os
import yaml

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

import gspread
from oauth2client.service_account import ServiceAccountCredentials

__author__ = 'sebastian'


# Loads some names from config.yaml
with open("config.yaml", 'r') as cfgyaml:
    try:
        cfg = yaml.load(cfgyaml)
    except yaml.YAMLError as exc:
        print(exc)

# Loads gspreadsheet ids from gspreadids.yaml
with open("config/gspreadids.yaml", 'r') as gsidsyaml:
    try:
        gsids = yaml.load(gsidsyaml)
    except yaml.YAMLError as exc:
        print(exc)

# Drive authorization
scope = ['https://spreadsheets.google.com/feeds']
key_filename = "key-for-gspread.json"
credentials = ServiceAccountCredentials.from_json_keyfile_name(key_filename, scope)
gc = gspread.authorize(credentials)


def get_sheetnames_by_date(filename, filter_key=""):
    wb = load_workbook(filename, read_only=True)
    sheetnames = [s for s in wb.sheetnames if filter_key in s]
    namesdates = [(name, load_tournament_xlsx(filename, name).date) for name in sheetnames]
    namesdates.sort(key=lambda p: p[1])

    return [name for name, date in namesdates]


def get_sheetnames_by_date_gs(spreadsheet_id, filter_key=""):
    wb = gc.open_by_key(spreadsheet_id)
    sheetnames = [s.title for s in wb.worksheets() if filter_key in s.title]
    namesdates = [(name, load_tournament_gs(spreadsheet_id, name).date) for name in sheetnames]
    namesdates.sort(key=lambda p: p[1])

    return [name for name, date in namesdates]


def load_sheet_workbook(filename, sheetname, first_row=1):
    wb = load_workbook(filename, read_only=True)
    ws = wb.get_sheet_by_name(sheetname)

    ws.calculate_dimension(force=True)
    # print(ws.dimensions)

    list_to_return = []
    max_column = 0
    for row in ws.rows:
        aux_row = []
        empty_row = True
        for cell in row:
            if cell.column:
                if cell.column > max_column:
                    max_column = cell.column
            if cell.value is None:
                aux_row.append("")
                # print(cell.column)
            else:
                empty_row = False
                aux_row.append(cell.value)
        if not empty_row:
            list_to_return.append(aux_row[:max_column])
    return list_to_return[first_row:]

# Dictionary to cache sheets on the same session and don't read them again
cache_sheets = dict()


def load_sheet_gs(spreadsheet_id, sheetname, first_row=1):
    """ Load a sheet and translated into a list.
    It won't read a sheet twice in the same session. They are chached on first read.
    """
    if (spreadsheet_id, sheetname) in cache_sheets:
        list_to_return = cache_sheets[(spreadsheet_id, sheetname)]
    else:
        wb = gc.open_by_key(spreadsheet_id)
        ws = wb.worksheet(sheetname)

        list_to_return = ws.get_all_values()  # Exclude null rows and columns

        # print(ws.row_count, ws.col_count)                   # Dimensions of sheet including null rows and columns
        # print(len(list_to_return), len(list_to_return[0]))  # Dimensions of sheet excluding null rows and columns

        cache_sheets[(spreadsheet_id, sheetname)] = list_to_return

    return list_to_return[first_row:]


def save_sheet_workbook(filename, sheetname, headers, list_to_save, overwrite=False):
    if os.path.isfile(filename):
        wb = load_workbook(filename)
        if overwrite and sheetname in wb:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        if sheetname in wb:
            ws = wb.get_sheet_by_name(sheetname)
        else:
            ws = wb.create_sheet()
    else:
        wb = Workbook()
        ws = wb.active

    ws.title = sheetname

    ws.append(headers)
    for col in range(1, ws.max_column+1):
        cell = ws.cell(column=col, row=1)
        cell.font = Font(bold=True)

    for row in list_to_save:
        ws.append(row)

    # # Automatically adjust width of columns to its content
    # # TODO add width adaptation, now it breaks on datetime
    # dims = {}
    # for row in ws.rows:
    #     for cell in row:
    #         if cell.value:
    #             dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    # for col, value in dims.items():
    #     ws.column_dimensions[col].width = value

    wb.save(filename)


def save_sheet_gs(spreadsheet_id, sheet_name, headers, rows_to_save):
    """ Saves headers and rows_to_save into given sheet_name.
        If sheet_name does not exist, it will be created. """

    wb = gc.open_by_key(spreadsheet_id)
    num_cols = len(headers)
    num_rows = len(rows_to_save) + 1  # +1 because of header

    # Overwrites an existing sheet or creates a new one
    if sheet_name in [ws.title for ws in wb.worksheets()]:
        ws = wb.worksheet(sheet_name)
        ws.resize(rows=num_rows, cols=num_cols)
    else:
        ws = wb.add_worksheet(title=sheet_name, rows=num_rows, cols=num_cols)
    # TODO create a new spreadsheet if it is not found

    # Concatenation of all cells values to be updated in batch mode
    cell_list = ws.range("A1:" + ws.get_addr_int(row=num_rows, col=num_cols))
    for i, value in enumerate(headers + [v for row in rows_to_save for v in row]):
        cell_list[i].value = value

    ws.update_cells(cell_list)

    # FIXME add bold to header
    # for col in range(1, len(ws.row_values())+1):
    #     cell = ws.cell(column=col, row=1)
    #     cell.font = Font(bold=True)

    # # Automatically adjust width of columns to its content
    # # TODO add width adaptation, now it breaks on datetime
    # dims = {}
    # for row in ws.rows:
    #     for cell in row:
    #         if cell.value:
    #             dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    # for col, value in dims.items():
    #     ws.column_dimensions[col].width = value


def save_ranking_sheet(filename, sheetname, ranking, players, overwrite=False):
    if os.path.isfile(filename):
        wb = load_workbook(filename)
        if overwrite and sheetname in wb:
            wb.remove_sheet(wb.get_sheet_by_name(sheetname))
        if sheetname in wb:
            ws = wb.get_sheet_by_name(sheetname)
        else:
            ws = wb.create_sheet()
    else:
        wb = Workbook()
        ws = wb.active

    ws.title = sheetname

    ws["A1"] = cfg["labels"]["Tournament name"]
    ws["B1"] = ranking.tournament_name
    ws.merge_cells('B1:G1')
    ws["A2"] = cfg["labels"]["Date"]
    ws["B2"] = ranking.date
    ws.merge_cells('B2:G2')
    ws["A3"] = cfg["labels"]["Location"]
    ws["B3"] = ranking.location
    ws.merge_cells('B3:G3')

    ws.append(cfg["labels"][key] for key in ["PID", "Total Points", "Rating Points", "Bonus Points",
                                             "Player", "Association", "City", "Active Player"])

    to_bold = ["A1", "A2", "A3",
               "A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4"]
    to_center = to_bold + ["B1", "B2", "B3"]

    for colrow in to_bold:
        cell = ws.cell(colrow)
        cell.font = Font(bold=True)
    for colrow in to_center:
        cell = ws.cell(colrow)
        cell.alignment = Alignment(horizontal='center')

    list_to_save = [[e.pid, e.get_total(), e.rating, e.bonus, players[e.pid].name, players[e.pid].association,
                     players[e.pid].city, str(ranking.tid - players[e.pid].last_tournament < 6)] for e in ranking]

    # for row in sorted(list_to_save, key=lambda l: (l[-1], l[1]), reverse=True):  # to use Jugador activo
    for row in sorted(list_to_save, key=lambda l: l[1], reverse=True):
        ws.append(row)

    wb.save(filename)


def save_ranking_sheet_gs(spreadsheet_id, sheet_name, ranking, players):
    """ Saves ranking into given sheet_name.
        If sheet_name does not exist, it will be created. """

    wb = gc.open_by_key(spreadsheet_id)

    headers = [cfg["labels"][key] for key in ["PID", "Total Points", "Rating Points", "Bonus Points",
                                              "Player", "Association", "City", "Active Player"]]
    rows_to_save = [[e.pid, e.get_total(), e.rating, e.bonus, players[e.pid].name, players[e.pid].association,
                     players[e.pid].city, str(ranking.tid - players[e.pid].last_tournament < 6)] for e in ranking]
    # for row in sorted(list_to_save, key=lambda l: (l[-1], l[1]), reverse=True):  # to use Jugador activo
    rows_to_save.sort(key=lambda l: l[1], reverse=True)

    num_cols = len(headers)
    num_rows = len(rows_to_save) + 1 + 3  # +1 because of header + 3 because of tournament metadata

    # Overwrites an existing sheet or creates a new one
    if sheet_name in [ws.title for ws in wb.worksheets()]:
        ws = wb.worksheet(sheet_name)
        ws.resize(rows=num_rows, cols=num_cols)
    else:
        ws = wb.add_worksheet(title=sheet_name, rows=num_rows, cols=num_cols)
    # TODO create a new spreadsheet if it is not found

    ws.update_acell("A1", "Nombre del torneo")
    ws.update_acell("B1", ranking.tournament_name)
    # TODO ws.merge_cells('B1:G1')
    ws.update_acell("A2", "Fecha")
    ws.update_acell("B2", ranking.date)
    # TODO ws.merge_cells('B2:G2')
    ws.update_acell("A3", "Lugar")
    ws.update_acell("B3", ranking.location)
    # TODO ws.merge_cells('B3:G3')

    # FIXME bold and center
    # to_bold = ["A1", "A2", "A3",
    #            "A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4"]
    # to_center = to_bold + ["B1", "B2", "B3"]
    #
    # for colrow in to_bold:
    #     cell = ws.cell(colrow)
    #     cell.font = Font(bold=True)
    # for colrow in to_center:
    #     cell = ws.cell(colrow)
    #     cell.alignment = Alignment(horizontal='center')

    # Concatenation of all cells values to be updated in batch mode
    cell_list = ws.range("A4:" + ws.get_addr_int(row=num_rows, col=num_cols))
    for i, value in enumerate(headers + [v for row in rows_to_save for v in row]):
        cell_list[i].value = value

    ws.update_cells(cell_list)


def load_ranking_sheet(filename, sheet_name):
    """Load a ranking in a xlxs sheet and return a Ranking object"""
    # TODO check if date is being read properly
    return load_ranking_list(load_sheet_workbook(filename, sheet_name, first_row=0))


def load_ranking_sheet_gs(spreadsheet_id, sheet_name):
    """Load a ranking in a spreadsheet sheet and return a Ranking object"""
    # TODO check if date is being read properly
    return load_ranking_list(load_sheet_gs(spreadsheet_id, sheet_name, first_row=0))


def load_ranking_list(raw_ranking_list):
    ranking = models.Ranking(raw_ranking_list[0][1], raw_ranking_list[1][1], raw_ranking_list[2][1])
    ranking.load_list([[int(rr[0]), int(rr[2]), int(rr[3])] for rr in raw_ranking_list[4:]])
    return ranking


def load_tournament_csv(filename):
    """Load a tournament csv and return a Tournament object"""
    with open(filename, 'r') as incsv:
        reader = csv.reader(incsv)
        tournament_list = [row for row in reader]
        return load_tournament_list(tournament_list)


def load_tournament_xlsx(filename, sheet_name):
    """Load a tournament xlsx sheet and return a Tournament object"""
    return load_tournament_list(load_sheet_workbook(filename, sheet_name, 0))


def load_tournament_gs(spreadsheet_id, sheet_name):
    """Load a tournament xlsx sheet and return a Tournament object"""
    return load_tournament_list(load_sheet_gs(spreadsheet_id, sheet_name, 0))


def load_tournament_list(tournament_list):
    """Load a tournament list sheet and return a Tournament object
    name = cell(B1)
    date = cell(B2)
    location = cell(B3)
    matches should be from sixth row containing:
    player1, player2, sets1, sets2, match_round, category
    """
    name = tournament_list[0][1]
    date = tournament_list[1][1]
    location = tournament_list[2][1]

    tournament = models.Tournament(name, date, location)

    # Reformated list of matches
    for player1, player2, sets1, sets2, round_match, category in tournament_list[5:]:
        # workaround to add extra bonus points from match list
        if int(sets1) < 0 and int(sets2) < 0:
            winner_name = cfg["aux"]["flag add bonus"]
            loser_name = player2
        elif int(sets1) > int(sets2):
            winner_name = player1
            loser_name = player2
        elif int(sets1) < int(sets2):
            winner_name = player2
            loser_name = player1
        else:
            print("Failed to process matches, a tie was found between %s and %s" % (player1, player2))
            break
        tournament.add_match(winner_name, loser_name, round_match, category)

    return tournament
